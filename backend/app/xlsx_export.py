"""Rebuild the full charted workbook straight from SQLite.

This is the one thing the browser cannot do: write a real .xlsx with native
chart objects. SheetJS in the frontend writes sheets and formulas but not charts.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
CATS = [
    ("Salary", "Income"), ("Other Income", "Income"),
    ("Rent / Housing", "Expense"), ("Groceries", "Expense"), ("Utilities", "Expense"),
    ("Internet & Phone", "Expense"), ("Transport", "Expense"), ("Dining Out", "Expense"),
    ("Health & Fitness", "Expense"), ("Insurance", "Expense"), ("Shopping", "Expense"),
    ("Entertainment", "Expense"), ("Subscriptions", "Expense"), ("Travel", "Expense"),
    ("Education", "Expense"), ("Gifts & Donations", "Expense"), ("Personal Care", "Expense"),
    ("Savings & Investments", "Expense"), ("Miscellaneous", "Expense"),
]
CUR = '"$"#,##0.00'
FONT = "Arial"
DARK, MID, RULE = "1F3B4D", "2E6E8E", "B7C9D3"
THIN = Side(style="thin", color=RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build(rows: list[dict], budget: dict, year: int, out_path: Path) -> Path:
    wb = Workbook()

    # ---- Transactions
    tx = wb.active
    tx.title = "Transactions"
    heads = ["Date", "Month", "Year", "Type", "Category", "Subcategory", "Description",
             "Amount (CAD)", "Payment Method", "Account", "Recurring?", "Notes"]
    tx.append(heads)
    _hdr(tx, 1, 1, len(heads))
    for r in sorted(rows, key=lambda x: str(x["date"])):
        d = str(r["date"])[:10]
        m = MONTHS[int(d[5:7]) - 1] if len(d) >= 7 else ""
        tx.append([d, m, int(d[:4]) if d[:4].isdigit() else "", r["type"], r["category"],
                   r["subcategory"], r["description"], float(r["amount"]),
                   r["payment"], r["account"], r["recurring"], r["notes"]])
    last = tx.max_row
    for row in tx.iter_rows(min_row=2, max_row=last, max_col=len(heads)):
        row[7].number_format = CUR
        for cell in row:
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
    for col, w in zip("ABCDEFGHIJKL", [12, 8, 8, 11, 22, 20, 30, 14, 20, 14, 12, 34]):
        tx.column_dimensions[col].width = w
    tx.freeze_panes = "A2"
    tx.auto_filter.ref = f"A1:L{max(last, 1)}"

    # ---- Budget
    bg = wb.create_sheet("Budget")
    bg.append(["Category", "Type", *MONTHS, "Annual Total"])
    _hdr(bg, 1, 1, 15)
    for i, (name, typ) in enumerate(CATS):
        r = 2 + i
        vals = [float(budget.get(name, {}).get(m, 0) or 0) for m in range(1, 13)]
        bg.append([name, typ, *vals, f"=SUM(C{r}:N{r})"])
    for row in bg.iter_rows(min_row=2, max_row=bg.max_row, max_col=15):
        for cell in row:
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
            if cell.column >= 3:
                cell.number_format = CUR
    bg.column_dimensions["A"].width = 24
    bg.freeze_panes = "C2"

    # ---- Pivot: live SUMIFS cross-tab, classic functions only
    pv = wb.create_sheet("Pivot")
    pv.append(["Category", "Type", *MONTHS, "Total"])
    _hdr(pv, 1, 1, 15)
    for i, (name, typ) in enumerate(CATS):
        r = 2 + i
        cells = [(f"=SUMIFS(Transactions!$H$2:$H${last},Transactions!$E$2:$E${last},$A{r},"
                  f"Transactions!$B$2:$B${last},{get_column_letter(3 + m)}$1)") for m in range(12)]
        pv.append([name, typ, *cells, f"=SUM(C{r}:N{r})"])
    tot = len(CATS) + 2
    pv.append(["TOTAL", "", *[f"=SUM({get_column_letter(3 + m)}2:{get_column_letter(3 + m)}{tot - 1})"
                              for m in range(12)], f"=SUM(C{tot}:N{tot})"])

    # KPI rows feed the charts
    k = tot + 2
    pv.cell(row=k, column=1, value="Total Income")
    pv.cell(row=k + 1, column=1, value="Total Expense")
    pv.cell(row=k + 2, column=1, value="Net Savings")
    pv.cell(row=k + 3, column=1, value="Expense Budget")
    for m in range(12):
        L = get_column_letter(3 + m)
        pv.cell(row=k, column=3 + m, value=f'=SUMIFS(Transactions!$H$2:$H${last},Transactions!$D$2:$D${last},"Income",Transactions!$B$2:$B${last},{L}$1)')
        pv.cell(row=k + 1, column=3 + m, value=f'=SUMIFS(Transactions!$H$2:$H${last},Transactions!$D$2:$D${last},"Expense",Transactions!$B$2:$B${last},{L}$1)')
        pv.cell(row=k + 2, column=3 + m, value=f"={L}{k}-{L}{k + 1}")
        pv.cell(row=k + 3, column=3 + m, value=f'=SUMIFS(Budget!{L}$2:{L}${len(CATS) + 1},Budget!$B$2:$B${len(CATS) + 1},"Expense")')
    for row in pv.iter_rows(min_row=2, max_row=k + 3, max_col=15):
        for cell in row:
            cell.border = BORDER
            if cell.column >= 3:
                cell.number_format = CUR
    pv.column_dimensions["A"].width = 24
    pv.freeze_panes = "C2"

    # ---- Dashboard with native charts
    db = wb.create_sheet("Dashboard")
    db["A1"] = f"EXPENSE DASHBOARD - {year} (CAD)"
    db["A1"].font = Font(name=FONT, bold=True, size=16, color=DARK)
    db["A3"] = "Charts read from the Pivot sheet. Regenerate this file from the app after adding data."
    db["A3"].font = Font(name=FONT, italic=True, size=9, color="666666")

    cats_ref = Reference(pv, min_col=3, max_col=14, min_row=1)

    c1 = BarChart(); c1.type = "col"; c1.title = f"Income vs Expense by Month - {year}"
    c1.add_data(Reference(pv, min_col=1, min_row=k, max_row=k + 1, max_col=14),
                titles_from_data=True, from_rows=True)
    c1.set_categories(cats_ref); c1.height, c1.width = 8.5, 17
    db.add_chart(c1, "A5")

    c2 = BarChart(); c2.type = "col"; c2.title = f"Net Savings by Month - {year}"
    c2.add_data(Reference(pv, min_col=1, min_row=k + 2, max_row=k + 2, max_col=14),
                titles_from_data=True, from_rows=True)
    c2.set_categories(cats_ref); c2.height, c2.width = 8.5, 17
    db.add_chart(c2, "J5")

    c3 = LineChart(); c3.title = "Actual Expense vs Budget Ceiling"
    c3.add_data(Reference(pv, min_col=1, min_row=k + 1, max_row=k + 1, max_col=14),
                titles_from_data=True, from_rows=True)
    c3.add_data(Reference(pv, min_col=1, min_row=k + 3, max_row=k + 3, max_col=14),
                titles_from_data=True, from_rows=True)
    c3.set_categories(cats_ref); c3.height, c3.width = 8.5, 17
    db.add_chart(c3, "A23")

    c4 = PieChart(); c4.title = f"Annual Spend by Category - {year}"
    c4.add_data(Reference(pv, min_col=15, min_row=4, max_row=len(CATS) + 1), titles_from_data=False)
    c4.set_categories(Reference(pv, min_col=1, min_row=4, max_row=len(CATS) + 1))
    c4.height, c4.width = 8.5, 17
    db.add_chart(c4, "J23")

    wb._sheets = [db, tx, bg, pv]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
